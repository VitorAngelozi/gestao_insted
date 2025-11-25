from django.shortcuts import render
from django.http import HttpResponse
from .models import Andar, Curso, Sala, SemestrePeriodo
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def homepage(request):
    # Buscar todas as salas que têm curso com semestre_periodo
    salas = Sala.objects.select_related('curso', 'andar', 'curso__semestre_periodo').filter(curso__semestre_periodo__isnull=False)
    
    # Filtros
    curso_id = request.GET.get('curso')
    andar_id = request.GET.get('andar')
    lugares_livres = request.GET.get('lugares_livres')
    semestre_id = request.GET.get('semestre')
    
    # Aplicar filtro por semestre (padrão: semestre ativo mais recente)
    if semestre_id:
        salas = salas.filter(curso__semestre_periodo_id=semestre_id)
    else:
        # Se não especificado, usar o semestre ativo mais recente
        semestre_ativo = SemestrePeriodo.objects.filter(ativo=True).order_by('-ano', '-periodo').first()
        if semestre_ativo:
            salas = salas.filter(curso__semestre_periodo=semestre_ativo)
    
    # Aplicar filtro por curso
    if curso_id:
        salas = salas.filter(curso_id=curso_id)
    
    # Aplicar filtro por andar
    if andar_id:
        salas = salas.filter(andar_id=andar_id)
    
    # Converter para lista para aplicar filtro de lugares livres
    salas_list = list(salas)
    
    # Aplicar filtro por lugares livres
    if lugares_livres == 'sim':
        salas_list = [sala for sala in salas_list if sala.tem_lugares_livres()]
    elif lugares_livres == 'nao':
        salas_list = [sala for sala in salas_list if not sala.tem_lugares_livres()]
    
    # Agrupar salas por andar
    andares_dict = {}
    for sala in salas_list:
        if sala.andar:  # Só processa se tiver andar
            andar_id = sala.andar_id
            if andar_id not in andares_dict:
                andares_dict[andar_id] = {
                    'andar': sala.andar,
                    'salas': []
                }
            andares_dict[andar_id]['salas'].append(sala)
    
    # Converter para lista e ordenar por número do andar
    andares_com_salas = sorted(andares_dict.values(), key=lambda x: x['andar'].numero if x['andar'] else 999)
    
    # Contexto para os filtros
    semestres = SemestrePeriodo.objects.all().order_by('-ano', '-periodo')
    semestre_selecionado = None
    if semestre_id:
        try:
            semestre_selecionado = SemestrePeriodo.objects.get(id=semestre_id)
        except SemestrePeriodo.DoesNotExist:
            pass
    elif not semestre_id:
        semestre_selecionado = SemestrePeriodo.objects.filter(ativo=True).order_by('-ano', '-periodo').first()
    
    # Filtrar cursos apenas do semestre selecionado para o dropdown
    if semestre_selecionado:
        cursos_filtrados = Curso.objects.filter(semestre_periodo=semestre_selecionado).order_by('nome')
    else:
        cursos_filtrados = Curso.objects.filter(semestre_periodo__isnull=False).order_by('nome')
    
    context = {
        'andares_com_salas': andares_com_salas,
        'cursos': cursos_filtrados,
        'andares': Andar.objects.all().order_by('numero'),
        'semestres': semestres,
        'semestre_selecionado': semestre_selecionado,
    }
    
    return render(request, 'homepage.html', context)

def exportar_excel(request):
    # Aplicar os mesmos filtros da homepage
    salas = Sala.objects.select_related('curso', 'andar', 'curso__semestre_periodo').filter(curso__semestre_periodo__isnull=False)
    
    curso_id = request.GET.get('curso')
    andar_id = request.GET.get('andar')
    lugares_livres = request.GET.get('lugares_livres')
    semestre_id = request.GET.get('semestre')
    
    if semestre_id:
        salas = salas.filter(curso__semestre_periodo_id=semestre_id)
    else:
        semestre_ativo = SemestrePeriodo.objects.filter(ativo=True).order_by('-ano', '-periodo').first()
        if semestre_ativo:
            salas = salas.filter(curso__semestre_periodo=semestre_ativo)
    
    if curso_id:
        salas = salas.filter(curso_id=curso_id)
    
    if andar_id:
        salas = salas.filter(andar_id=andar_id)
    
    salas_list = list(salas)
    
    if lugares_livres == 'sim':
        salas_list = [sala for sala in salas_list if sala.tem_lugares_livres()]
    elif lugares_livres == 'nao':
        salas_list = [sala for sala in salas_list if not sala.tem_lugares_livres()]
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salas"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # Título
    ws['A1'] = "Relatório de Salas - Faculdade Insted"
    ws['A1'].font = title_font
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Data de exportação
    ws['A2'] = f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws.merge_cells('A2:H2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Cabeçalhos
    headers = ['Andar', 'Sala', 'Curso', 'Semestre Curso', 'Turma', 'Alunos', 'Lugares', 'Disponíveis', 'Status']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados
    row_num = 5
    for sala in salas_list:
        ws.cell(row=row_num, column=1).value = sala.andar.numero if sala.andar else "N/A"
        ws.cell(row=row_num, column=2).value = sala.nome
        ws.cell(row=row_num, column=3).value = sala.curso.nome
        ws.cell(row=row_num, column=4).value = sala.curso.get_semestre_display()
        ws.cell(row=row_num, column=5).value = sala.curso.turma
        ws.cell(row=row_num, column=6).value = sala.curso.alunos
        ws.cell(row=row_num, column=7).value = sala.lugares
        ws.cell(row=row_num, column=8).value = sala.lugares_disponiveis()
        ws.cell(row=row_num, column=9).value = "Disponível" if sala.tem_lugares_livres() else "Lotado"
        
        # Formatação condicional para status
        status_cell = ws.cell(row=row_num, column=9)
        if sala.tem_lugares_livres():
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        row_num += 1
    
    # Ajustar largura das colunas
    column_widths = [12, 20, 30, 15, 10, 10, 10, 12, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"relatorio_salas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
